#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.

# TODO: refactor into smaller functions

import openpyxl
from openpyxl.styles.borders import Border, Side

import datetime
import os

import config.srrs_config as config
import config.products as products

import lib_srrs_util as util


def main():
    year = config.target_year

    sales_dir = os.path.join(config.sales_dir_fqn, str(year))

    #calculate_variants = True
    calculate_variants = False

    # maps from the VAT percentage to what actually needs to be used in the calculation
    vat_map = { 0.06: 0.0566, 0.12: 0.1071, 0.25: 0.20}
    
    out_wb, out_ws = create_out_wb_and_ws() 
    
    out_row_count = 2

    total_commission_sales = 0
    first_product = True
    
    input_files = [f for f in os.listdir(sales_dir) if os.path.isfile(os.path.join(sales_dir, f))]
    print("\nSearching for sales in: " + str(input_files))
    
    target_month = int(input("\nEnter the month number: "))
    
    for i in range(0, len(input_files)):
        input_files[i] = os.path.join(sales_dir, input_files[i])
    
    days = set()
    rest_products = set()
   
    bersa_sum = {k: 0 for k in products.bersa}
    bersa_count = bersa_sum.copy()
    
    fika_sum = {k: 0 for k in products.fika}
    fika_count = fika_sum.copy()
    
    hyllhyra_sum = {k: 0 for k in products.hyllhyra}
    hyllhyra_count = hyllhyra_sum.copy()
    
    print ("\nLoading sales data...\n")
    # holds the sheets with the sales data
    sheets = []
    
    for input_file in input_files:
        wb = openpyxl.load_workbook(filename = input_file, data_only=True)
        sheets.append(wb['Sheet 1'])
    
    
    print ("Summing sales...\n")
    
    for product in products.product_categories:
        item_sum = 0
        item_count = 0
    
        # row A = date, row B = time, row E = product, row F = variant, row H = amount, row K = price
        for sheet in sheets:
            # iterate over all cells that contain data
            for row in range(config.sales_first_data_row, sheet.max_row + 1):
                date_str = sheet["A" + str(row)].value
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        
                # check if the current row is within the given time window
                if (util.is_target_month(date, target_month, days)):
                    cur_product = sheet["E" + str(row)].value
                    
                    if (cur_product in products.product_categories): # own products sold
                        if (product == cur_product):
                            variant = sheet["F" + str(row)].value
                            amount = sheet["H" + str(row)].value
                            price = sheet["K" + str(row)].value
        
                            item_count = item_count + amount
                            item_sum = item_sum + price
        
                            if calculate_variants:
                                if product == products.bersa_category:
                                    bersa_sum[variant] = bersa_sum[variant] + price
                                    bersa_count[variant] = bersa_count[variant] + amount
                                if product == products.fika_category:
                                    fika_sum[variant] = fika_sum[variant] + price
                                    fika_count[variant] = fika_count[variant] + amount
                                if product == products.hyllhyra_category:
                                    hyllhyra_sum[variant] = hyllhyra_sum[variant] + price
                                    hyllhyra_count[variant] = hyllhyra_count[variant] + amount
                        
                    else: # products sold by others, calculate commission
                        rest_products.add(cur_product)
                        if not first_product or cur_product in products.no_commission:
                            continue
                        price = sheet["K" + str(row)].value
                        total_commission_sales = total_commission_sales + price
    
                        #cur_date = sheet["A" + str(row)].value
                        #cur_time = sheet["B" + str(row)].value
                        #print("%s, %s: Adding price %d for product %s to total_commission_sales." % (cur_date, cur_time, price, cur_product))
        
        # computations are done here
        # writing results to excel sheet (out_ws)
        
        print ("In month {} product {:20s} sold {:5d} times for: {:5d} sek.".format(
            target_month, product, int(item_count), int(item_sum)))
        
        if product == products.fika_category: # food has 12% VAT
            VAT = 0.12
        elif product == "Book": # books have 6% VAT
            VAT = 0.06
        else: # everything else has 25% VAT
            VAT = 0.25
    
        if (product == products.fika_category or product == products.kh_category
            or product == products.hyllhyra_category):
            out_row_count = out_row_count + 1 # skip a row to have a space between categories?
    
        # A = Produkt, B = Variant, C = Items, D = Sales, E = VAT, G = Sales without VAT
        if product == products.kh_category:
            out_ws["A" + str(out_row_count)] = "KH - Begangnade varor"
        else:
            out_ws["A" + str(out_row_count)] = product
        out_ws["C" + str(out_row_count)] = int(item_count)
        out_ws["D" + str(out_row_count)] = int(item_sum)
        out_ws["E" + str(out_row_count)] = "=ROUND(D" + str(out_row_count) + "*" + str(vat_map[VAT]) + ")"
        out_ws["F" + str(out_row_count)] = VAT
        out_ws["G" + str(out_row_count)] = "=D" + str(out_row_count) + "-" + "E" + str(out_row_count)
        out_row_count = out_row_count + 1
    
        if calculate_variants:
            if product == products.bersa_category:
                for var in bersa_count:
                    print("\t{} variant {:20s} sold {:5d} items for {:5d} sek.".format(
                        products.bersa_category, var,  bersa_count[var], bersa_sum[var]))
    
                    out_ws["B" + str(out_row_count)] = var
                    out_ws["C" + str(out_row_count)] = int(bersa_count[var])
                    out_ws["D" + str(out_row_count)] = int(bersa_sum[var])
                    out_row_count = out_row_count + 1
    
            if product == products.fika_category:
                for var in fika_count:
                    print("\t{} variant {:20s} sold {:5d} items for {:5d} sek.".format(
                        products.fika_category, var, fika_count[var], fika_sum[var]))
    
                    out_ws["B" + str(out_row_count)] = var
                    out_ws["C" + str(out_row_count)] = int(fika_count[var])
                    out_ws["D" + str(out_row_count)] = int(fika_sum[var])
                    out_row_count = out_row_count + 1
    
            if product == products.hyllhyra_category:
                for var in hyllhyra_count:
                    print("\t{} variant {:20s} sold {:5d} items for {:5d} sek.".format(
                        products.hyllhyra_category, var,
                        hyllhyra_count[var], hyllhyra_sum[var]))
    
                    out_ws["B" + str(out_row_count)] = var
                    out_ws["C" + str(out_row_count)] = int(hyllhyra_count[var])
                    out_ws["D" + str(out_row_count)] = int(hyllhyra_sum[var])
                    out_row_count = out_row_count + 1
    
        first_product = False
    
    
    # output which days were part of the calculation (mostly good for debugging)
    #out_row_count = out_row_count + 1
    #out_ws["A" + str(out_row_count)] = ""
    #print("\nSales were found on the following days in month %d:" % target_month)
    #for d in days:
    #    print(d, end=", ")
    #    out_ws["A" + str(out_row_count)] = out_ws["A" + str(out_row_count)].value + "," + d
    #print("\n")
    
    #print ("\nProducts not covered in calculation, but found in sales:\n", str(rest_products))
    #for x in no_commission:
    #    rest_products.discard(x)
    #print ("\nProducts to calculate commission from:\n", str(rest_products))
    
    print ("\nTotal commission sales:", total_commission_sales)
    print ("Total commission:", total_commission_sales * 0.2)
    
    # add commision
    #out_row_count = out_row_count + 1
    out_ws["A" + str(out_row_count)] = "Kommission (Hyllor)"
    out_ws["D" + str(out_row_count)] = round( int(total_commission_sales) * 0.2)
    out_ws["E" + str(out_row_count)] = "=ROUNd(D" + str(out_row_count) + "*" + str(vat_map[0.25]) + ")"
    out_ws["F" + str(out_row_count)] = 0.25
    out_ws["G" + str(out_row_count)] = "=D" + str(out_row_count) + "-" + "E" + str(out_row_count)
    
    # add total sum
    out_row_count = out_row_count + 2
    
    out_ws["A" + str(out_row_count)] = "Summa"
    out_ws["D" + str(out_row_count)] = "=SUM(D2:D" + str(out_row_count - 2) + ")"
    out_ws["E" + str(out_row_count)] = "=SUM(E2:E" + str(out_row_count - 2) + ")"
    out_ws["G" + str(out_row_count)] = "=SUM(G2:G" + str(out_row_count - 2) + ")"
    
    top_border = Border(top=Side(style='thin'))
    out_ws["A" + str(out_row_count)].border = top_border
    out_ws["B" + str(out_row_count)].border = top_border
    out_ws["C" + str(out_row_count)].border = top_border
    out_ws["D" + str(out_row_count)].border = top_border
    out_ws["E" + str(out_row_count)].border = top_border
    out_ws["F" + str(out_row_count)].border = top_border
    out_ws["G" + str(out_row_count)].border = top_border
    
    
    out_row_count = out_row_count + 2
    out_ws["A" + str(out_row_count)] = "This file was autogenerated: CHANGES WILL BE OVERWRITTEN -> Make a copy for changes!"
    
    out_file = os.path.join(sales_dir, "kh-monthly-summaries", "%d-%s-sales-summary.xlsx" % (year,'{:02d}'.format(target_month)))
    out_wb.save(out_file)
    
    util.pause()

    
def create_out_wb_and_ws():
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    
    out_ws["A1"] = "Produkt"
    out_ws["B1"] = "Variant"
    out_ws["C1"] = "Antal"
    out_ws["D1"] = "Brutto"
    out_ws["E1"] = "Varav moms"
    out_ws["F1"] = "Moms (%)"
    out_ws["G1"] = "Netto"
    
    out_ws.column_dimensions["A"].width = 20
    out_ws.column_dimensions["B"].width = 20
    out_ws.column_dimensions["C"].width = 7
    out_ws.column_dimensions["E"].width = 13
    #out_ws.column_dimensions["G"].width = 15
    
    #ft = Font(color=colors.RED)
    ft = openpyxl.styles.Font(bold=True)
    out_ws["A1"].font = out_ws["B1"].font = out_ws["C1"].font = out_ws["D1"].font = out_ws["E1"].font = out_ws["F1"].font = out_ws["G1"].font = ft
    
    bottom_border = Border(bottom=Side(style='thin'))
    out_ws["A1"].border = out_ws["B1"].border = out_ws["C1"].border = out_ws["D1"].border = out_ws["E1"].border = out_ws["F1"].border = out_ws["G1"].border = bottom_border

    return (out_wb, out_ws)


if __name__ == '__main__':
    main()
    
