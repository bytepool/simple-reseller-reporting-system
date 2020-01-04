#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.


import openpyxl
import datetime
import math
import os

import config.srrs_config as config
import lib_srrs_util as util


def locate_input_files():
    sales_dir = os.path.join(config.sales_dir_fqn, str(config.target_year))
    input_files = [f for f in os.listdir(sales_dir) if os.path.isfile(os.path.join(sales_dir, f))]

    print("\nSearching for sales in: " + str(input_files))

    # append directory to file names
    for i in range(0, len(input_files)):
        input_files[i] = os.path.join(sales_dir, input_files[i])

    return input_files


def load_sales_data(input_files):
    print("\nLoading sales data...")
    sheets = []

    for input_file in input_files:
        wb = openpyxl.load_workbook(filename = input_file, data_only=True)
        sheets.append(wb['Sheet 1'])
    
    print("\nSales data successfully loaded.")
    return sheets


def gen_shelf_strings():
    shelfs = []
    
    # generate shelf strings from H01 to H34
    for i in range(1,35):
        shelf = "H" + '{:02d}'.format(int(i))
        shelfs.append(shelf)
            
    # generate special shelf strings from S01 to S20
    for i in range(1,21):
        shelf = "S" + '{:02d}'.format(int(i))
        shelfs.append(shelf)

    return shelfs


def create_new_worksheet(out_wb):
    out_ws = out_wb.active
    
    out_ws["A1"] = "Shelf"
    out_ws["B1"] = "Items"
    out_ws["C1"] = "Sales"
    out_ws["D1"] = "Payout"

    return out_ws


def sum_sales(shelfs, input_sheets, target_week, days):
    print ("\nSumming sales...\n")

    sales = []
    
    for shelf in shelfs:
        item_sum = 0
        item_count = 0
    
        # row A = date, row B = time, row E = product,
        # row F = variant, row H = amount, row K = price
        for sheet in input_sheets:
            # iterate over all cells that contain data
            for row in range(config.sales_first_data_row, sheet.max_row + 1):
                date_str = sheet["A" + str(row)].value
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        
                # check if the current row is within the given time window
                if (util.is_target_week(date, target_week, days)):
                    product = sheet["E" + str(row)].value
                    if (shelf in product):
                        variant = sheet["F" + str(row)].value
                        amount = sheet["H" + str(row)].value
                        price = sheet["K" + str(row)].value

                        item_count = item_count + amount
                        item_sum = item_sum + price

        sales.append((item_count, item_sum))

    print("\nSales data successfully calculated.")
    return sales
    

def write_sales(shelfs, target_week, sales, out_ws, out_wb):
    for i in range(0, len(shelfs)):
        row_index = i + 2
        shelf = shelfs[i]
        item_count = sales[i][0]
        item_sum = sales[i][1]
        
        print ("In week {} shelf {:10} sold {:4d} items for {:6d} "
               "sek".format(target_week, shelf, int(item_count), int(item_sum)))
    
        out_ws["A" + str(row_index)] = shelf
        out_ws["B" + str(row_index)] = item_count
        out_ws["C" + str(row_index)] = item_sum
        
        payout = item_sum - round(item_sum * 0.2)
        out_ws["D" + str(row_index)] = payout

    sales_dir = os.path.join(config.sales_dir_fqn, str(config.target_year))
    out_file = os.path.join(sales_dir, "weekly-summaries", "%s-v%s-sales-summary.xlsx" % (str(config.target_year), '{:02d}'.format(int(target_week))))
    
    out_wb.save(out_file)
    print("\nData successfully written to {:s}".format(out_file))

    
def calc_and_write_sales_summary():
    days = set()

    input_files = locate_input_files()
    input_sheets = load_sales_data(input_files)
    shelfs = gen_shelf_strings()

    target_week = util.ask_target_week(str(config.target_year))

    out_wb = openpyxl.Workbook()
    out_ws = create_new_worksheet(out_wb)

    sales = sum_sales(shelfs, input_sheets, target_week, days)

    write_sales(shelfs, target_week, sales, out_ws, out_wb)
    
    print("\nSanity check:\nSales were found on the following days in week %d:" % target_week)
    for d in days:
        print(d, end=", ")
    print("\n")

    util.pause()

    
if __name__ == '__main__':
    calc_and_write_sales_summary()
    
