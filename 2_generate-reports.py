#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.


import os
import re
import glob
import datetime
import openpyxl
import subprocess

from os import path

from lib_srrs_customer_db import CustomerEntry, get_customer_entry
from lib_srrs_bookings_db import BookingsWorkbook, BookingEntry

import config.srrs_config as config
import lib_srrs_util as util


def create_target_dir(target_week):
    report_dir = path.join(config.report_dir_fqn, str(config.target_year))
    if not path.exists(report_dir):
        os.mkdir(report_dir)
        
    wd = path.join(report_dir, "vecka{:02d}".format(int(target_week)))
    
    if not path.exists(wd):
        os.mkdir(wd)

    return wd
 

def read_template_str():
    try:
        template_fh = open(config.report_template_file_fqn, 'r')
        template_str = template_fh.read()
        template_fh.close()
    except Exception as e:
        print("ERROR: Could not open or read template file. Exception: {:s}".format(str(e)))
        util.pause()
        exit(1)

    return template_str


def create_tex_filename(shelf, target_week):
    tex_filename = "kundrapport-{:s}-v{:02d}.tex".format(shelf, int(target_week))
    return tex_filename


def write_file(fn, wd, content):
    fqn = path.join(wd, fn)
    try:
        tex_file = open(fqn, 'w')
        tex_file.write(content)
        tex_file.close()
    except Exception as e:
        print("ERROR: Could not write file {:s}: {:s}".format(fn, str(e)))
        util.pause()
        exit(1)


def remove_tmp_files(wd):
    to_remove = []
    to_remove.extend(glob.glob(wd + path.sep + "*.aux"))
    to_remove.extend(glob.glob(wd + path.sep + "*.log"))
    to_remove.extend(glob.glob(wd + path.sep + "*.out"))

    for filename in to_remove:
        os.remove(filename)

    
def instantiate_report(template, customer, sales, target_week):
    report = re.sub(r'--WEEKNR--', "{:02d}".format(int(target_week)), template)

    # converting strings with str() is necessary in case the variable is of type None
    report = re.sub(r'--FULLNAME--', str(customer.name), report)
    report = re.sub(r'--ADDRESS--', str(customer.address), report)
    report = re.sub(r'--PHONE--', str(customer.phone), report)
    report = re.sub(r'--EMAIL--', str(customer.email), report)
    report = re.sub(r'--CUSTOMERNR--', str(customer.customer_nr), report)
    report = re.sub(r'--ACCOUNTNR--', str(customer.account_nr), report)

    report = re.sub(r'--NAME--', str(sales.customer_first_name), report)
    report = re.sub(r'--SALES--', str(sales.sales), report)
    report = re.sub(r'--ITEMS--', str(sales.nr_items), report)
    report = re.sub(r'--COMISSION--', str(sales.commission), report)
    report = re.sub(r'--PAYOUT--', str(sales.pay_out), report)
    return report


def replace_umlaute(in_str):
    out = re.sub(re.compile("ä"), r'{\"a}', in_str)
    out = re.sub(re.compile("ö"), r'{\"o}', out)
    out = re.sub(re.compile("ü"), r'{\"u}', out)
    out = re.sub(re.compile("å"), r'{\\aa}', out)
    out = re.sub(re.compile("Ä"), r'{\"A}', out)
    out = re.sub(re.compile("Ö"), r'{\"O}', out)
    out = re.sub(re.compile("Ü"), r'{\"U}', out)
    out = re.sub(re.compile("Å"), r'{\\AA}', out)
    out = re.sub(re.compile("_"), r'\_', out)
    out = re.sub(re.compile("é"), r'\'{e}', out)
    return out


def main():
    target_week = util.ask_target_week(str(config.target_year))
    
    working_dir = create_target_dir(target_week)
    
    # load data 
    bw = BookingsWorkbook(config.bookings_file_fqn, data_only=True)
    bookings_sheet = bw.getBookingSheetOfWeekNr(target_week)

    wb = openpyxl.load_workbook(filename = config.customer_reg_file_fqn, data_only=True)
    customers_sheet = wb["Kundregister"]

    template_str = read_template_str()
    
    # iterate over all shelfs in the bookings file
    for i in range(config.bookings_first_row, bookings_sheet.max_row + 1):
        try:
            booking_entry = BookingEntry(bookings_sheet, i)
        except Exception as e:
            print("WARNING: Could not parse row {:02d}: {:s}.".format(i, str(e)))
            continue

        shelf = booking_entry.shelf
        
        if not booking_entry.customer_nr:
            print ("Skipping shelf {:s} due to empty customer ID.".format(shelf))
            continue        

        customer_entry = get_customer_entry(customers_sheet, booking_entry.customer_nr)

        report = instantiate_report(template_str, customer_entry, booking_entry, target_week)
        report = replace_umlaute(report)

        tex_filename = create_tex_filename(shelf, target_week)
        write_file(fn=tex_filename, wd=working_dir, content=report)

        print("Creating report for shelf {:s}...".format(shelf))

        try:
            subprocess.check_call([config.latex_cmd, config.latex_options, tex_filename],
                                  cwd=working_dir, stdout=subprocess.DEVNULL)
        except subprocess.CalledProcessError as e:
            print("\nERROR: Could not create report for shelf {:s}!"
                  "Check {:s} for errors.\n".format(shelf, tex_filename))
            continue

    print ("\nSuccessfully created reports in {:s}.".format(working_dir))
    
    remove_tmp_files(working_dir)
    print ("Removed all temporary files in {:s}.\n".format(working_dir))

    util.pause()

    
if __name__ == '__main__':
    main()

