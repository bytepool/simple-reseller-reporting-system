#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.


import os
import time
import smtplib
import imaplib
import ssl
import getpass
import openpyxl

from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.utils import formataddr
from email.header import Header

from lib_srrs_bookings_db import BookingsWorkbook, BookingEntry
from lib_srrs_customer_db import CustomerEntry, get_customer_entry

import config.srrs_config as config
import lib_srrs_util as util


def confirm_sending(week_nr, year):
    confirm = input("Are you sure you want to send the emails for week "
                    "{:d} ({:s}) (y/n)? ".format(week_nr, year))

    if (confirm == "y") or (confirm == "yes"):
        return
    else:
        print("Aborting program as requested.")
        util.pause()
        exit(0)
    

def main():

    year = str(config.target_year)
    
    bookings_file = config.bookings_file_fqn
    customer_file = config.customer_reg_file_fqn

    week_nr = util.ask_target_week(year)

    confirm_sending(week_nr, year)
    
    bw = BookingsWorkbook(bookings_file, data_only=True)
    bookings_sheet = bw.getBookingSheetOfWeekNr(week_nr)

    wb = openpyxl.load_workbook(filename = customer_file, data_only=True)
    customer_reg_sheet = wb["Kundregister"]

    server = connect_to_server()

    # iterate over all shelfs in bookings file
    for i in range(config.bookings_first_row, bookings_sheet.max_row + 1):
        try:
            booking_entry = BookingEntry(bookings_sheet, i)
        except Exception as e:
            print("Unknown Error: skipping shelf in row %d." % i)
            continue

        shelf = booking_entry.shelf
        
        if not booking_entry.customer_nr:
            print ("Skipping shelf {:s} due to empty customer ID.".format(shelf))
            continue

        customer_entry = get_customer_entry(customer_reg_sheet, booking_entry.customer_nr)

        msg = create_msg(shelf, customer_entry.name, customer_entry.email, week_nr, year)
        if msg:
            #time.sleep(1) # wait one second, to avoid being labelled as spam
            server.send_message(msg)
            print("OK: Message for shelf {:s} sent.".format(shelf))
        else:
            print("Failure: Message for shelf {:s} could not be sent! Check "
                  "email address and report for potential errors.".format(shelf))

    server.quit()

    print("All emails that could be send have been sent.")
    util.pause()

    
def create_msg(shelf, name, email, week_nr, year):
    #print ("trying to create msg for", name, email)
    if not email:
        return None

    sender = formataddr((str(Header(u'Kvillehyllan', 'utf-8')), "kontakt@kvillehyllan.se"))

    #print("DEBUG: report_dir_fqn = %s, year = %s, week_str = %s" % (config.report_dir_fqn, year, week_str))
    report_filedir = os.path.join(config.report_dir_fqn, str(year),
                                  "vecka{:02d}".format(week_nr))
    template_file = config.email_template_file_fqn
    
    try:
        receiver = formataddr((str(Header(name, 'utf-8')), email))
    except UnicodeEncodeError: #email address is not in ASCII
        return None

    msg = MIMEMultipart()
    msg['Subject'] = config.email_subject % ("{:02d}".format(week_nr), shelf)
    msg['From'] = sender
    msg['BCC'] = sender
    msg['To'] = receiver

    report_filename = "kundrapport-{:s}-v{:02d}.pdf".format(shelf, week_nr)

    try:
        with open(os.path.join(report_filedir, report_filename), 'rb') as report:
            pdf = MIMEApplication(report.read(), _subtype='pdf')
            pdf.add_header('content-disposition', 'attachment', filename=report_filename)
            msg.attach(pdf)
    except FileNotFoundError:
        return None

    with open(template_file, encoding='utf-8') as template:
        # Create a text/plain message from template_file
        mime_text = MIMEText(template.read())

    msg.attach(mime_text)
    return msg

def connect_to_server():
    smtp_server = config.smtp_server
    smtp_port = config.smtp_port

    login_name = config.email_login_name
    
    try:
        server = smtplib.SMTP(host=smtp_server, port=smtp_port)

        # TODO: Check if the context bug has been fixed by now, see below. 
        # Without a standard context, the connection will NOT be encrypted. 
        # However, there is a bug which currently stops this from working,
        # so we are not using it right now.
        #context = ssl.create_default_context()
        
        print("Connecting to {:s} on port {:d}...".format(smtp_server, smtp_port))
        server.connect(host=smtp_server, port=smtp_port)

        print("Sending ehlo...")
        server.ehlo()

        #print("Sending starttls with context %s..." % str(context))
        #server.starttls(context=context)
        print("Sending starttls without context...")
        server.starttls()
        
        passw = getpass.getpass("Enter password: ")
        print("Logging in...")
        
        server.login(login_name, passw)
        
    except Exception as e:
        print("An unexpected error occurred while connecting to the server:", e.args)
        util.pause()

    return server


if __name__ == '__main__':
    main()


