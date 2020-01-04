#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.

# TODO: refactor into smaller functions

import openpyxl
from openpyxl.styles.borders import Border, Side

import datetime
import os

import config.srrs_config as config
import config.products as products

import lib_srrs_util as util


def create_out_wb_and_ws(columns):
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    
    out_ws["A1"] = "Produkt"
    out_ws["A2"] = "Brutto"
    out_ws["A3"] = "Moms 25%"
    out_ws["A4"] = "Moms 12%"
    out_ws["A5"] = "Moms  6%"
    out_ws["A6"] = "Netto"
    
    for i in range(0,len(columns)):
        out_ws.column_dimensions[columns[i]].width = 15
    
    #ft = Font(color=colors.RED)
    ft = openpyxl.styles.Font(bold=True)
    out_ws["A1"].font = out_ws["A2"].font = out_ws["A3"].font = out_ws["A4"].font = out_ws["A5"].font = out_ws["A6"].font = ft

    return (out_wb, out_ws)


def main():
    year = config.target_year
    sales_dir = os.path.join(config.sales_dir_fqn, str(year))
    
    
    # Make room for 26 products
    columns = [chr(a) for a in range(ord("A"), ord("Z") + 1)]
    product_count = 1
    
    # maps from the VAT percentage to what actually needs to be used in the calculation
    vat_map = { 0.06: 0.0566, 0.12: 0.1071, 0.25: 0.20}
    
    out_wb, out_ws = create_out_wb_and_ws(columns)
    
    total_commission_sales = 0
    first_product = True
    
    input_files = [f for f in os.listdir(sales_dir) if os.path.isfile(os.path.join(sales_dir, f))]
    print("\nSearching for sales in: " + str(input_files))
    
    for i in range(0, len(input_files)):
        input_files[i] = os.path.join(sales_dir, input_files[i])

    target_quarter = int(input("\nEnter the quarter number (1-4): "))
    
    days = set()
    rest_products = set()
    sheets = []

    print("\nLoading sales data...")
    
    for input_file in input_files:
        wb = openpyxl.load_workbook(filename = input_file, data_only=True)
        sheets.append(wb['Sheet 1'])
    
    print("\nSumming sales...\n")
    
    for product in products.product_categories:
        item_sum = 0
    
        # row A = date, row B = time, row E = product, row F = variant, row H = amount, row K = price
        for sheet in sheets:
            # iterate over all cells that contain data
            for row in range(config.sales_first_data_row, sheet.max_row + 1):
                date_str = sheet["A" + str(row)].value
                date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        
                # check if the current row is within the given time window
                if (util.is_target_quarter(date, target_quarter, days)):
                    cur_product = sheet["E" + str(row)].value
                    
                    if (cur_product in products): # products sold by Kvillehyllan
                        if (product == cur_product):
                            amount = sheet["H" + str(row)].value
                            price = sheet["K" + str(row)].value
        
                            item_sum = item_sum + price
                        
                    else: # products sold by others, calculate commission
                        rest_products.add(cur_product)
                        if not first_product or cur_product in no_commission:
                            continue
                        price = sheet["K" + str(row)].value
                        total_commission_sales = total_commission_sales + price
    
                        #cur_date = sheet["A" + str(row)].value
                        #cur_time = sheet["B" + str(row)].value
                        #print("%s, %s: Adding price %d for product %s to total_commission_sales." % (cur_date, cur_time, price, cur_product))
        
        # computations are done here
        # writing results to excel sheet (out_ws)
        
        print ("In Q{} product {:20s} sold for: {:5d} sek.".format(
            target_quarter, product, int(item_sum)))
        
        if product == products.fika_category:
            VAT = 0.12
        elif product == "Book":
            VAT = 0.06
        else:
            VAT = 0.25
    
        # A1 = Produkt, A2 = Brutto, A3 = Moms 25%, A4 = Moms 12%, A5 = Moms 6%, A6 = Netto
        if product == products.kh_category:
            out_ws[columns[product_count] + "1"] = "KH - Begangnad"
        else:
            out_ws[columns[product_count] + "1"] = product
    
        out_ws[columns[product_count] + "2"] = int(item_sum)
        vat_index = ""
        if VAT == 0.06:
            vat_index = columns[product_count] + "5"
            out_ws[vat_index] = "=" + columns[product_count] + "2" + "*" + str(vat_map[VAT]) + ""
        elif VAT == 0.12:
            vat_index = columns[product_count] + "4"
            out_ws[vat_index] = "=" + columns[product_count] + "2" + "*" + str(vat_map[VAT]) + ""
        elif VAT == 0.25:
            vat_index = columns[product_count] + "3"
            out_ws[vat_index] = "=" + columns[product_count] + "2" + "*" + str(vat_map[VAT]) + ""
            
        out_ws[columns[product_count] + "6"] = "=" + columns[product_count] + "2" + "-" + vat_index
        product_count = product_count + 1
        first_product = False
    
    
    # output which days were part of the calculation (mostly good for debugging)
    #product_count = product_count + 1
    #out_ws["A" + str(product_count)] = ""
    #print("\nSales were found on the following days in Q%d:" % target_quarter)
    #for d in days:
    #    print(d, end=", ")
    #    out_ws["A" + str(product_count)] = out_ws["A" + str(product_count)].value + "," + d
    #print("\n")
    
    #print ("\nProducts not covered in calculation, but found in sales:\n", str(rest_products))
    #for x in no_commission:
    #    rest_products.discard(x)
    #print ("\nProducts to calculate commission from:\n", str(rest_products))
    
    print ("\nTotal commission sales:", total_commission_sales)
    print ("Total commission:", total_commission_sales * 0.2)
    
    # add commision
    out_ws[columns[product_count] + "1"] = "Kommission"
    out_ws[columns[product_count] + "2"] = round(int(total_commission_sales) * 0.2)
    out_ws[columns[product_count] + "3"] = "=" + columns[product_count] + "2" + "*" + str(vat_map[0.25])
    out_ws[columns[product_count] + "6"] = "=" + columns[product_count] + "2" + "-" + columns[product_count] + "3"
    
    
    ## add total sum
    product_count = product_count + 1
    out_ws[columns[product_count] + "1"] = "Summa"
    out_ws[columns[product_count] + "2"] = "=SUM(B2:" + columns[product_count - 1] + "2)"
    out_ws[columns[product_count] + "3"] = "=SUM(B3:" + columns[product_count - 1] + "3)"
    out_ws[columns[product_count] + "4"] = "=SUM(B4:" + columns[product_count - 1] + "4)"
    out_ws[columns[product_count] + "5"] = "=SUM(B5:" + columns[product_count - 1] + "5)"
    out_ws[columns[product_count] + "6"] = "=SUM(B6:" + columns[product_count - 1] + "6)"
    
    top_border = Border(top=Side(style='thin'))
    for i in range(0,len(products)):
        out_ws[columns[i] + "7"].border = top_border
    
    bottom_border = Border(bottom=Side(style='thin'))
    for i in range(0,len(products)):
        out_ws[columns[i] + "1"].border = bottom_border
        
    out_ws["A9"] = "This file was autogenerated: CHANGES WILL BE OVERWRITTEN!"
    
    out_file = os.path.join(sales_dir, "kh-quarterly-summaries", "%d-Q%s-sales-summary.xlsx" % (year,target_quarter))
    out_wb.save(out_file)
    
    util.pause()

        
if __name__ == '__main__':
    main()

    
