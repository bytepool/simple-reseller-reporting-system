#!/usr/bin/python3

# simple reseller reporting system
# 
# Copyright (C) 2016 - 2019, Aljoscha Lautenbach. 
#
# This program is free software; you can redistribute it
# and/or modify it under the terms of the GNU General
# Public License version 2 as published by the Free
# Software Foundation
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public
# License along with this program; if not, write to the
# Free Software Foundation, Inc., 51 Franklin Street,
# Fifth Floor, Boston, MA 02110-1301 USA.


import openpyxl
import datetime
import math
import os

import config.srrs_config as config

new_year = 2020
input_file = os.path.join(config.bookings_dir_fqn, "bookings-{}.xlsx".format(new_year))

first_new_week = 1
last_week = 52

#wb = openpyxl.load_workbook(filename = input_file, data_only=True) # formula evaluation is used
wb = openpyxl.load_workbook(filename = input_file) # fields with formulas contain the formula strings, not the evaluated values


def get_first_day_of_first_week():
    day = 1
    date = datetime.datetime(new_year, 1, day)
    while date.isocalendar()[1] != 1:
        day = day + 1
        date = datetime.datetime(new_year, 1, day)
    return date


def get_first_day_of_week(weeknr):
    first_day = get_first_day_of_first_week()
    delta = datetime.timedelta(weeks=(weeknr - 1))
    return first_day + delta


def main():
    # iterate over all chosen weeks
    week_delta = datetime.timedelta(days=6)
    
    for week in range(first_new_week, last_week + 1):
        next_sheet_str = 'V' + '{:02d}'.format(week)
        current_sheet_str = 'V' + '{:02d}'.format(week - 1)

        if not next_sheet_str in wb:
            print ("Creating sheet " + next_sheet_str)
            next_sheet = wb.create_sheet(title=next_sheet_str)
        else:
            print ("Selecting sheet " + next_sheet_str)
            next_sheet = wb[next_sheet_str]    
        
        current_sheet = wb[current_sheet_str]

        print ("Copying cells for new sheet " + next_sheet_str)

        for r in current_sheet.iter_rows(current_sheet.calculate_dimension()):
            for cell in r:
                new_cell = next_sheet[cell.coordinate]
                new_cell.data_type = cell.data_type
                new_cell.style = cell.style
                new_cell.value = cell.value
                
        date = get_first_day_of_week(week)
        first_day = date.strftime("%a %Y-%m-%d")
        last_day = (date + week_delta).strftime("%a %Y-%m-%d")
        next_sheet['A1'] = first_day + " - " + last_day
        next_sheet['A2'] = "Vecka " + '{:02d}'.format(week)
        
    wb.save(input_file)
    print("Finished creating new weeks.")
    
def is_target_week(date):
    week = date.isocalendar()[1]
    if  (week == target_week):
        days.add(date.strftime("%a %Y-%m-%d"))
        return True
    else:
        return False

    
if __name__ == '__main__':
    main()


